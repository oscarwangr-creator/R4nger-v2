"""
RedTeam Framework v2.0 - Module: osint/identity_fusion
Flagship SOCMINT Identity Investigation Pipeline — Claude AI Enhanced

Architecture:
  Stage A   — Seed normalization & entity graph init
  Stage B   — Username sweep (14+ tools, parallel)
  Stage B2  — ★ Deep account scraping (platform-specific + generic HTTP)
  Stage B3  — ★ Multi-engine web search (Google, Bing, Brave, DDG, Yahoo, StartPage)
  Stage C   — Email pivot & breach intelligence
  Stage D   — Social media deep dive
  Stage E   — Domain & name intelligence
  Stage F   — Code/secrets/git forensics
  Stage G   — Phone OSINT
  Stage H   — Dark web & breach data
  Stage I   — Geolocation, metadata, wayback
  Stage J   — Claude AI correlation & confidence scoring
  Stage K   — Multi-format report output

Stage B2 takes every URL discovered in Stage B and scrapes it:
  • Platform-specific tools for Twitter/X, Instagram, GitHub, Reddit,
    LinkedIn, TikTok, YouTube, Twitch, Steam, Keybase, Mastodon, etc.
  • Generic async HTTP scraper (httpx) for any remaining URL
  • Bio/description, follower counts, following counts, linked accounts,
    emails, phone numbers, location, website, join date, post samples
  • All harvested entities flow back into the main investigation graph

Tools universe: 90+ SOCMINT tools with graceful degradation.
Each stage skips missing tools and continues with available ones.
"""

from __future__ import annotations

import asyncio
import csv
import json
import os
import re
import shutil
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import urlparse

from framework.modules.base import BaseModule, Finding, ModuleResult, Severity

# ─────────────────────────────────────────────────────────────────────────────
# Regex helpers
# ─────────────────────────────────────────────────────────────────────────────

EMAIL_RE   = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
URL_RE     = re.compile(r"https?://[^\s\])\>\"']+")
PHONE_RE   = re.compile(r"\+?[\d\-\(\)\s]{8,20}")
USERNAME_RE= re.compile(r"\b[a-zA-Z0-9._\-]{3,32}\b")

# ─────────────────────────────────────────────────────────────────────────────
# Platform detection from URL
# ─────────────────────────────────────────────────────────────────────────────

# Maps hostname fragment → (platform_name, scraper_type)
PLATFORM_MAP: Dict[str, Tuple[str, str]] = {
    "twitter.com":      ("twitter",    "twint"),
    "x.com":            ("twitter",    "twint"),
    "instagram.com":    ("instagram",  "instaloader"),
    "github.com":       ("github",     "gitfive"),
    "reddit.com":       ("reddit",     "http"),
    "tiktok.com":       ("tiktok",     "http"),
    "youtube.com":      ("youtube",    "http"),
    "twitch.tv":        ("twitch",     "http"),
    "steam":            ("steam",      "http"),
    "keybase.io":       ("keybase",    "http"),
    "mastodon":         ("mastodon",   "http"),
    "linkedin.com":     ("linkedin",   "http"),
    "facebook.com":     ("facebook",   "http"),
    "pinterest.com":    ("pinterest",  "http"),
    "tumblr.com":       ("tumblr",     "http"),
    "flickr.com":       ("flickr",     "http"),
    "deviantart.com":   ("deviantart", "http"),
    "patreon.com":      ("patreon",    "http"),
    "medium.com":       ("medium",     "http"),
    "gitlab.com":       ("gitlab",     "http"),
    "bitbucket.org":    ("bitbucket",  "http"),
    "soundcloud.com":   ("soundcloud", "http"),
    "spotify.com":      ("spotify",    "http"),
    "last.fm":          ("lastfm",     "http"),
    "discord":          ("discord",    "http"),
    "telegram":         ("telegram",   "http"),
    "onlyfans.com":     ("onlyfans",   "http"),
    "substack.com":     ("substack",   "http"),
    "hackernews":       ("hackernews", "http"),
    "lobste.rs":        ("lobsters",   "http"),
    "producthunt.com":  ("producthunt","http"),
    "quora.com":        ("quora",      "http"),
    "stackoverflow.com":("stackoverflow","http"),
    "behance.net":      ("behance",    "http"),
    "dribbble.com":     ("dribbble",   "http"),
    "vimeo.com":        ("vimeo",      "http"),
    "dailymotion.com":  ("dailymotion","http"),
}

# ─────────────────────────────────────────────────────────────────────────────
# Stage B — Username sweep tools
# ─────────────────────────────────────────────────────────────────────────────

STAGE_B_USERNAME_TOOLS = [
    ("sherlock",        "sherlock",       ["sherlock", "{username}", "--print-found", "--timeout", "10"],            "300+ social networks"),
    ("maigret",         "maigret",        ["maigret", "{username}", "--no-color", "--timeout", "30"],                "OSINT by username"),
    ("nexfil",          "nexfil",         ["nexfil", "-u", "{username}"],                                           "OSINT username finder"),
    ("blackbird",       "blackbird",      ["blackbird", "-u", "{username}"],                                        "Account OSINT"),
    ("socialscan",      "socialscan",     ["socialscan", "{username}", "--json"],                                    "Username/email check"),
    ("whatsmyname",     "wmn",            ["wmn", "-u", "{username}"],                                              "Username lookup"),
    ("social-analyzer", "social-analyzer",["social-analyzer", "--username", "{username}", "--silent"],              "Social analytics"),
    ("snoop",           "snoop",          ["python3", "-m", "snoop", "{username}"],                                 "Username OSINT"),
    ("username-anarchy","username-anarchy",["username-anarchy", "{username}"],                                      "Username variations"),
    ("osrframework",    "usufy",          ["usufy", "-n", "{username}"],                                            "OSRFramework users"),
    ("seekr",           "seekr",          ["seekr", "user", "{username}"],                                          "OSINT search tool"),
    ("profil3r",        "profil3r",       ["profil3r", "{username}"],                                               "Profile OSINT"),
    ("namechk",         "namechk",        ["namechk", "{username}"],                                               "Username availability"),
    ("checkusernames",  "checkusernames", ["checkusernames", "{username}"],                                         "Username checker"),
    ("peekyou",         "peekyou-cli",    ["peekyou-cli", "--username", "{username}"],                             "PeekYou search"),
    ("usersearch",      "usersearch",     ["usersearch", "{username}"],                                            "UserSearch.org"),
    ("sherlock-project","sherlock",       ["python3", "-m", "sherlock_project", "{username}"],                     "Sherlock pip variant"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Stage B2 — Platform-specific scraper commands
# ─────────────────────────────────────────────────────────────────────────────

# Maps scraper_type → command template (supports {username}, {url}, {platform_user})
STAGE_B2_SCRAPER_COMMANDS: Dict[str, List[str]] = {
    "twint":       ["twint", "-u", "{username}", "--profile-full",
                    "-o", "/tmp/twint_{username}.json", "--json"],
    "instaloader": ["instaloader", "--no-pictures", "--no-videos",
                    "--no-metadata-json", "{username}"],
    "gitfive":     ["gitfive", "user", "{username}"],
    "toutatis":    ["toutatis", "-u", "{username}"],
    "osintgram":   ["python3", "main.py", "{username}"],
    "twscrape":    ["twscrape", "user_by_login", "{username}"],
    "snscrape":    ["snscrape", "--jsonl", "--max-results", "50",
                    "twitter-user", "{username}"],
}

# ─────────────────────────────────────────────────────────────────────────────
# Stage C — Email tools
# ─────────────────────────────────────────────────────────────────────────────

STAGE_C_EMAIL_TOOLS = [
    ("holehe",        "holehe",      ["holehe", "{email}", "--only-used"],                           "Email account checker"),
    ("h8mail",        "h8mail",      ["h8mail", "-t", "{email}", "-q"],                             "Email breach hunter"),
    ("emailrep",      "emailrep",    ["emailrep", "{email}"],                                       "Email reputation"),
    ("ignorant",      "ignorant",    ["ignorant", "{email}"],                                       "Ignorant email check"),
    ("dehashed-cli",  "dehashed",    ["dehashed", "search", "{email}"],                             "DeHashed breach search"),
    ("breach-parse",  "breach-parse",["breach-parse", "{email}"],                                   "Breach data parser"),
    ("emailfinder",   "emailfinder", ["emailfinder", "-d", "{domain}"],                            "Email finder"),
    ("haveibeenpwned","hibp",        ["hibp", "{email}"],                                           "HIBP check"),
]

STAGE_D_SOCIAL_TOOLS = [
    ("instaloader",     "instaloader",   ["instaloader", "--no-pictures", "--no-videos", "{username}"],   "Instagram OSINT"),
    ("toutatis",        "toutatis",      ["toutatis", "-u", "{username}"],                               "Instagram account info"),
    ("twint",           "twint",         ["twint", "-u", "{username}", "--profile-full", "-o", "/tmp/twint_out.json", "--json"], "Twitter OSINT"),
    ("tinfoleak",       "tinfoleak",     ["python3", "tinfoleak.py", "-u", "{username}"],                "Twitter info leak"),
    ("socid-extractor", "socid_extractor",["socid_extractor", "--input", "{username}"],                  "Social ID extractor"),
    ("osintgram",       "osintgram",     ["python3", "main.py", "{username}"],                          "Instagram geo OSINT"),
    ("twscrape",        "twscrape",      ["twscrape", "user_by_login", "{username}"],                   "Twitter scraper"),
    ("snscrape",        "snscrape",      ["snscrape", "--jsonl", "twitter-user", "{username}"],         "Social network scrape"),
    ("reddit-user-analyser","reddit-user-analyser",["reddit-user-analyser", "{username}"],              "Reddit user analysis"),
]

STAGE_E_DOMAIN_TOOLS = [
    ("theHarvester", "theHarvester", ["theHarvester", "-d", "{domain}", "-b", "all", "-l", "500", "-f", "/tmp/harvest_{domain}"], "Email/subdomain harvest"),
    ("spiderfoot",   "spiderfoot",   ["spiderfoot", "-s", "{domain}", "-m", "all", "-q"],             "OSINT automation"),
    ("metagoofil",   "metagoofil",   ["metagoofil", "-d", "{domain}", "-t", "pdf,doc,xls,ppt", "-o", "/tmp/meta_{domain}"], "Metadata extraction"),
    ("dnstwist",     "dnstwist",     ["dnstwist", "--mxcheck", "--whois", "{domain}"],               "Domain squatting"),
    ("linkedin2username","linkedin2username",["python3", "linkedin2username.py", "{domain}"],          "LinkedIn username gen"),
    ("crosslinked",  "crosslinked",  ["crosslinked", "{domain}"],                                    "LinkedIn name enum"),
    ("photon",       "photon",       ["python3", "photon.py", "-u", "https://{domain}", "--quiet"],   "Web crawler OSINT"),
    ("recon-ng",     "recon-ng",     ["recon-ng", "-h"],                                             "Recon framework"),
    ("subfinder",    "subfinder",    ["subfinder", "-d", "{domain}", "-silent", "-all"],              "Subfinder (recon)"),
    ("amass",        "amass",        ["amass", "intel", "-d", "{domain}", "-whois"],                  "AMASS intelligence"),
    ("shodan",       "shodan",       ["shodan", "host", "{domain}"],                                 "Shodan lookup"),
    ("censys",       "censys",       ["censys", "search", "{domain}"],                               "Censys.io search"),
    ("urlscan",      "urlscan",      ["urlscan", "search", "{domain}"],                              "URLScan.io"),
]

STAGE_F_CODE_TOOLS = [
    ("gitfive",   "gitfive",   ["gitfive", "user", "{username}"],                                 "GitHub OSINT"),
    ("octosuite", "octosuite", ["octosuite", "--username", "{username}"],                         "GitHub suite OSINT"),
    ("trufflehog","trufflehog",["trufflehog", "github", "--repo", "https://github.com/{username}"],"Secret scanner"),
    ("gitleaks",  "gitleaks",  ["gitleaks", "detect", "--no-git", "--report-format", "json", "--report-path", "/tmp/gitleaks.json"], "Git secrets"),
    ("gitrecon",  "gitrecon",  ["python3", "gitrecon.py", "-u", "{username}"],                   "Git recon"),
    ("gharchive", "gharchive", ["gharchive", "{username}"],                                      "GitHub archive data"),
]

STAGE_G_PHONE_TOOLS = [
    ("phoneinfoga",      "phoneinfoga",      ["phoneinfoga", "scan", "-n", "{phone}"],            "Phone OSINT"),
    ("email2phonenumber","email2phonenumber",["email2phonenumber", "{email}"],                    "Email→phone pivot"),
]

STAGE_H_DARKWEB_TOOLS = [
    ("intelx",    "intelx",  ["intelx", "-search", "{username}"],   "IntelX search"),
    ("pwndb-cli", "pwndb",   ["pwndb", "--query", "{email}"],       "PwnDB breach lookup"),
    ("dehashed",  "dehashed",["dehashed", "search", "{username}"],  "DeHashed search"),
    ("skymem",    "skymem",  ["skymem", "{email}"],                "Skymem email exposure"),
]

STAGE_I_GEO_TOOLS = [
    ("creepy",         "creepy",         ["creepy", "-t", "{username}"],                                            "Geolocation OSINT"),
    ("waybackpack",    "waybackpack",    ["waybackpack", "https://{domain}", "-d", "/tmp/wayback_{domain}"],         "Wayback Machine pack"),
    ("ghunt",          "ghunt",          ["ghunt", "account", "{email}"],                                          "Google account OSINT"),
    ("sn0int",         "sn0int",         ["sn0int", "run", "--quiet"],                                             "Sn0int framework"),
    ("buster",         "buster",         ["python3", "buster.py", "{email}"],                                      "Profile finder"),
    ("exiftool",       "exiftool",       ["exiftool", "-json"],                                                    "Metadata from images"),
    ("recon-ng",       "recon-ng",       ["recon-ng", "-h"],                                                       "Recon-ng framework"),
    ("archivebox",     "archivebox",     ["archivebox", "add", "https://{domain}"],                                "Archive snapshot"),
    ("harpoon",        "harpoon",        ["harpoon", "help"],                                                      "OSINT/CTI tool"),
    ("waymore",        "waymore",        ["waymore", "-i", "{domain}", "-oR", "/tmp/waymore_{domain}"],            "Waymore recon"),
    ("tineye",         "tineye",         ["tineye", "{image_url}"],                                               "Reverse image TinEye"),
    ("social-searcher","social-searcher",["social-searcher", "{username}"],                                        "Social searcher"),
    ("mentionmapp",    "mentionmapp",    ["mentionmapp", "{username}"],                                            "Twitter mention map"),
    ("clearbit",       "clearbit",       ["clearbit", "--email", "{email}"],                                      "Clearbit enrichment"),
    ("fullcontact",    "fullcontact",    ["fullcontact", "--email", "{email}"],                                    "FullContact lookup"),
    ("pipl",           "pipl",           ["pipl", "--email", "{email}"],                                          "Pipl people search"),
    ("hunter",         "hunter",         ["hunter", "domain", "{domain}"],                                        "Hunter.io email finder"),
    ("phantombuster",  "phantombuster",  ["phantombuster", "--help"],                                             "PhantomBuster automations"),
]

ADDITIONAL_SOCMINT_TOOLS = [
    "snscrape", "twscrape", "social-searcher", "reddit-user-analyser", "wikileaks-search",
    "archivebox", "waymore", "email2phonenumber", "fullcontact", "clearbit", "hunter",
    "skymem", "pipl", "onyphe", "zoomeye", "criminalip", "fullhunt", "netlas", "wigle",
    "spyonweb", "dnsdumpster", "urlscan", "intelowl", "amass", "gau", "subfinder",
    "dehashed", "foca", "tineye", "mentionmapp", "netlytic", "peekyou", "knowem",
    "usersearch", "exiftool", "phantombuster", "shlink", "censys", "shodan", "gharchive",
    "harpoon", "cybelangel", "maltego", "recon-ng", "photon", "crosslinked", "linkedin2username",
]


# ─────────────────────────────────────────────────────────────────────────────
# Scraped profile dataclass
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ScrapedProfile:
    url: str
    platform: str
    username: str = ""
    display_name: str = ""
    bio: str = ""
    followers: Optional[int] = None
    following: Optional[int] = None
    posts: Optional[int] = None
    location: str = ""
    website: str = ""
    joined: str = ""
    verified: bool = False
    emails_found: List[str] = field(default_factory=list)
    phones_found: List[str] = field(default_factory=list)
    linked_accounts: List[str] = field(default_factory=list)
    linked_urls: List[str] = field(default_factory=list)
    post_samples: List[str] = field(default_factory=list)
    raw_text: str = ""
    scrape_method: str = "http"
    scrape_success: bool = False
    error: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "url": self.url, "platform": self.platform,
            "username": self.username, "display_name": self.display_name,
            "bio": self.bio[:400] if self.bio else "",
            "followers": self.followers, "following": self.following,
            "posts": self.posts, "location": self.location,
            "website": self.website, "joined": self.joined,
            "verified": self.verified,
            "emails_found": self.emails_found,
            "phones_found": self.phones_found,
            "linked_accounts": self.linked_accounts,
            "linked_urls": self.linked_urls[:20],
            "post_samples": self.post_samples[:5],
            "scrape_method": self.scrape_method,
            "scrape_success": self.scrape_success,
        }


# ─────────────────────────────────────────────────────────────────────────────
# HTML extraction patterns per platform
# ─────────────────────────────────────────────────────────────────────────────

# Regex patterns to pull profile data from raw HTML responses
_HTML_PATTERNS: Dict[str, List[Tuple[str, str]]] = {
    "twitter": [
        ("bio",       r'"description":\s*"([^"]{5,500})"'),
        ("followers", r'"followers_count":\s*(\d+)'),
        ("following", r'"friends_count":\s*(\d+)'),
        ("posts",     r'"statuses_count":\s*(\d+)'),
        ("location",  r'"location":\s*"([^"]{2,100})"'),
        ("website",   r'"url":\s*"(https?://[^"]+)"'),
        ("joined",    r'"created_at":\s*"([^"]+)"'),
        ("name",      r'"name":\s*"([^"]{1,100})"'),
    ],
    "instagram": [
        ("bio",       r'"biography":\s*"([^"]{5,500})"'),
        ("followers", r'"edge_followed_by":\s*\{[^}]*"count":\s*(\d+)'),
        ("following", r'"edge_follow":\s*\{[^}]*"count":\s*(\d+)'),
        ("posts",     r'"edge_owner_to_timeline_media":\s*\{[^}]*"count":\s*(\d+)'),
        ("website",   r'"external_url":\s*"([^"]+)"'),
        ("name",      r'"full_name":\s*"([^"]{1,100})"'),
    ],
    "github": [
        ("bio",       r'class="p-note[^"]*"[^>]*>([^<]{5,500})<'),
        ("followers", r'(\d+)\s+followers'),
        ("following", r'(\d+)\s+following'),
        ("repos",     r'(\d+)\s+repositories'),
        ("location",  r'class="p-label[^"]*"[^>]*>([^<]{2,100})<'),
        ("website",   r'class="Link--primary[^"]*"\s+href="(https?://[^"]+)"'),
        ("joined",    r'Joined on\s*<[^>]+>([^<]+)<'),
        ("name",      r'class="p-name[^"]*"[^>]*>([^<]{1,100})<'),
    ],
    "reddit": [
        ("bio",       r'"subreddit":\s*\{[^}]*"public_description":\s*"([^"]{5,300})"'),
        ("followers", r'"icon_img".*?"total_karma":\s*(\d+)'),
        ("joined",    r'"created":\s*(\d+)'),
        ("name",      r'"name":\s*"([^"]{1,50})"'),
    ],
    "default": [
        ("bio",       r'<meta[^>]+(?:name|property)=["\']description["\'][^>]+content=["\']([^"\']{10,400})["\']'),
        ("bio2",      r'<meta[^>]+content=["\']([^"\']{10,400})["\'][^>]+(?:name|property)=["\']description["\']'),
        ("name",      r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']{1,100})["\']'),
        ("image",     r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']'),
        ("website",   r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)["\']'),
    ],
}


def _detect_platform(url: str) -> Tuple[str, str]:
    """Return (platform_name, scraper_type) from a URL."""
    try:
        host = urlparse(url).netloc.lower().replace("www.", "")
        for fragment, (platform, stype) in PLATFORM_MAP.items():
            if fragment in host:
                return platform, stype
    except Exception:
        pass
    return "unknown", "http"


def _extract_username_from_url(url: str, platform: str) -> str:
    """Best-effort extract a username from a profile URL."""
    try:
        path = urlparse(url).path.strip("/")
        parts = [p for p in path.split("/") if p]
        if not parts:
            return ""
        # For platforms where first path segment is the username
        skip = {"user", "u", "profile", "in", "people", "members", "channel"}
        if parts[0].lower() in skip and len(parts) > 1:
            return parts[1]
        return parts[0]
    except Exception:
        return ""


class IdentityFusionModule(BaseModule):
    """
    9-stage + Stage B2 SOCMINT identity investigation with Claude AI.

    Stage B2 scrapes every discovered account URL using:
      1. Platform-specific CLI tools (twint, instaloader, gitfive, etc.)
      2. Generic async httpx scraping with per-platform HTML extraction
      3. JSON/API endpoint probing for structured profile data
    All scraped entities (emails, URLs, names, linked accounts) flow
    back into the investigation entity graph for further pivoting.
    """

    def info(self) -> Dict[str, Any]:
        return {
            "name": "identity_fusion",
            "description": (
                "11-stage SOCMINT investigation (90+ tools) + Stage B2 deep account scraping "
                "+ Stage B3 multi-engine web search (Google/Bing/Brave/DDG/Yahoo/StartPage). "
                "Discovers accounts → scrapes profiles → searches the web for mentions "
                "→ pivots on all entities → Claude AI correlation → multi-format report."
            ),
            "author": "RTF Core Team",
            "category": "osint",
            "version": "4.0",
            "references": [
                "https://github.com/sherlock-project/sherlock",
                "https://github.com/soxoj/maigret",
                "https://github.com/megadose/holehe",
                "https://github.com/smicallef/spiderfoot",
            ],
            "tags": ["osint", "socmint", "identity", "investigation", "scraping"],
        }

    def _declare_options(self) -> None:
        self._register_option("username",       "Seed username",                                   required=False, default="")
        self._register_option("email",          "Seed email address",                              required=False, default="")
        self._register_option("full_name",      "Seed full name",                                  required=False, default="")
        self._register_option("phone",          "Seed phone number (E.164 format)",                required=False, default="")
        self._register_option("domain",         "Seed/pivot domain",                               required=False, default="")
        self._register_option("image_url",      "URL of target image for reverse search",          required=False, default="")
        self._register_option("timeout",        "Per-tool timeout seconds",                        required=False, default=180, type=int)
        self._register_option("scrape_timeout", "Per-URL HTTP scrape timeout seconds",             required=False, default=20,  type=int)
        self._register_option("max_scrape_urls","Max account URLs to scrape in Stage B2",          required=False, default=50,  type=int)
        self._register_option("scrape_accounts","Enable Stage B2 deep account scraping",           required=False, default=True, type=bool)
        self._register_option("scrape_js",      "Use JS rendering (requires playwright)",          required=False, default=False,type=bool)
        self._register_option("max_pivots",     "Max deduped items per entity type",               required=False, default=500, type=int)
        self._register_option("concurrency",    "Parallel tools per stage",                        required=False, default=4,   type=int)
        self._register_option("scrape_concurrency","Parallel scraper connections",                 required=False, default=8,   type=int)
        self._register_option("stages",         "Comma-separated stages B-K or 'all'",             required=False, default="all")
        self._register_option("tool_profile",   "core|full|aggressive",                            required=False, default="core", choices=["core", "full", "aggressive"])
        self._register_option("output_format",  "json|csv|xlsx|pdf|html",                          required=False, default="json", choices=["json", "csv", "xlsx", "pdf", "html"])
        self._register_option("output_file",    "Explicit output file path",                       required=False, default="")
        self._register_option("use_ai",         "Use Claude AI for correlation stage J",           required=False, default=True, type=bool)
        self._register_option("shodan_key",     "Shodan API key",                                  required=False, default="")
        self._register_option("intelx_key",     "IntelX API key",                                  required=False, default="")
        self._register_option("dehashed_key",   "DeHashed API key",                                required=False, default="")
        # Stage B3 — Multi-engine web search
        self._register_option("web_search",     "Enable Stage B3 multi-engine web search",         required=False, default=True, type=bool)
        self._register_option("search_engines", "Engines: duckduckgo,bing,brave,google,yahoo,startpage",
                              required=False, default="duckduckgo,bing,brave")
        self._register_option("search_results", "Results per search engine per query",             required=False, default=10, type=int)
        self._register_option("search_fetch",   "Scrape each search result page",                  required=False, default=True, type=bool)
        self._register_option("search_max_fetch","Max result pages to scrape per query",           required=False, default=5, type=int)
        self._register_option("search_delay",   "Delay between engine requests (seconds)",         required=False, default=1.5, type=float)
        self._register_option("google_api_key", "Google Custom Search API key (optional)",         required=False, default="")
        self._register_option("google_cx",      "Google Custom Search CX ID (optional)",           required=False, default="")
        self._register_option("brave_api_key",  "Brave Search API key (optional)",                 required=False, default="")

    def validate(self) -> None:
        super().validate()
        seeds = [self.get(k) for k in ("username","email","full_name","phone","domain")]
        if not any(seeds):
            raise ValueError("At least one seed required: username, email, full_name, phone, or domain")

    # ─────────────────────────────────────────────────────────────────
    # Main run()
    # ─────────────────────────────────────────────────────────────────

    async def run(self) -> ModuleResult:
        start_ts = datetime.utcnow()
        timeout          = self.get("timeout")
        max_pivots       = self.get("max_pivots")
        concurrency      = self.get("concurrency")
        profile          = self.get("tool_profile")
        stages_raw       = self.get("stages")
        active_stages    = set("BCDEFGHIJK") if stages_raw == "all" else set(stages_raw.upper().split(","))
        do_scrape        = self.get("scrape_accounts")
        max_scrape_urls  = self.get("max_scrape_urls")
        scrape_timeout   = self.get("scrape_timeout")
        scrape_conc      = self.get("scrape_concurrency")
        domain           = self.get("domain") or ""

        usernames: Set[str] = {self.get("username")} if self.get("username") else set()
        emails:    Set[str] = {self.get("email")}    if self.get("email")    else set()
        names:     Set[str] = {self.get("full_name")} if self.get("full_name") else set()
        phones:    Set[str] = {self.get("phone")}    if self.get("phone")    else set()
        urls:      Set[str] = set()

        tool_runs:     List[Dict[str, Any]] = []
        events:        List[Dict[str, Any]] = []
        stage_results: Dict[str, Any]       = {}
        scraped_profiles: List[Dict[str,Any]] = []

        def _tokens() -> Dict[str, str]:
            return {
                "username":  sorted(usernames)[0] if usernames else "",
                "email":     sorted(emails)[0]    if emails    else "",
                "full_name": sorted(names)[0]     if names     else "",
                "domain":    domain,
                "phone":     sorted(phones)[0]    if phones    else "",
                "image_url": self.get("image_url") or "",
            }

        # ── Stage B: Username sweep ─────────────────────────────────
        if "B" in active_stages and usernames:
            self.log.info(f"[B] Username sweep — {len(STAGE_B_USERNAME_TOOLS)} tools")
            b_res = await self._run_stage("B", STAGE_B_USERNAME_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(b_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(b_res["tool_runs"]); events.extend(b_res["events"])
            stage_results["B"] = b_res["summary"]

        # ── Stage B2: Deep account scraping ─────────────────────────
        if "B" in active_stages and do_scrape and urls:
            self.log.info(f"[B2] Deep account scraping — {min(len(urls), max_scrape_urls)} URLs")
            b2_profiles, b2_new = await self._scrape_found_accounts(
                account_urls=sorted(urls)[:max_scrape_urls],
                seed_username=sorted(usernames)[0] if usernames else "",
                timeout=scrape_timeout,
                concurrency=scrape_conc,
                profile=profile,
            )
            scraped_profiles.extend(b2_profiles)
            # Harvest new entities from scraped data
            for prof in b2_profiles:
                if prof.get("scrape_success"):
                    emails.update(    e for e in prof.get("emails_found",   []) if e)
                    urls.update(      u for u in prof.get("linked_urls",    []) if u)
                    names_from = [prof.get("display_name","")]
                    names.update(     n for n in names_from if n)
                    linked = prof.get("linked_accounts", [])
                    usernames.update( u for u in linked if u)
            stage_results["B2"] = {
                "profiles_scraped": len(b2_profiles),
                "profiles_succeeded": sum(1 for p in b2_profiles if p.get("scrape_success")),
                "new_emails": len(b2_new.get("emails",[])),
                "new_urls": len(b2_new.get("urls",[])),
                "new_accounts": len(b2_new.get("usernames",[])),
            }
            self.log.info(f"[B2] Scraped {stage_results['B2']['profiles_succeeded']}/{stage_results['B2']['profiles_scraped']} profiles — "
                          f"+{stage_results['B2']['new_emails']} emails, +{stage_results['B2']['new_accounts']} accounts")

        # ── Stage B3: Multi-engine web search ────────────────────
        do_web_search = self.get("web_search")
        if "B" in active_stages and do_web_search:
            search_queries = []
            if self.get("username"): search_queries.append(f'"{self.get("username")}"')
            if self.get("full_name"): search_queries.append(f'"{self.get("full_name")}"')
            if self.get("email"): search_queries.append(self.get("email"))
            if domain: search_queries.append(f'site:{domain}')
            # Add cross-platform pivot queries
            if self.get("username"):
                un = self.get("username")
                search_queries += [
                    f'"{un}" site:twitter.com OR site:instagram.com OR site:github.com',
                    f'"{un}" email OR contact OR profile',
                ]
            if search_queries:
                self.log.info(f"[B3] Web search — {len(search_queries)} queries × {self.get('search_engines')}")
                b3_results, b3_emails, b3_urls, b3_social = await self._run_web_search_stage(
                    queries=search_queries[:4],  # limit to 4 queries to avoid hammering
                    engines=self.get("search_engines"),
                    results_per_engine=self.get("search_results"),
                    fetch_pages=self.get("search_fetch"),
                    max_fetch=self.get("search_max_fetch"),
                    timeout=self.get("scrape_timeout"),
                    delay=self.get("search_delay"),
                    google_key=self.get("google_api_key"),
                    google_cx=self.get("google_cx"),
                    brave_key=self.get("brave_api_key"),
                )
                emails.update(b3_emails)
                urls.update(b3_urls)
                stage_results["B3"] = {
                    "queries_run": len(search_queries[:4]),
                    "total_search_results": len(b3_results),
                    "new_emails": len(b3_emails),
                    "new_urls": len(b3_urls),
                    "social_links": len(b3_social),
                }
                self.log.info(f"[B3] {stage_results['B3']['total_search_results']} search results | "
                              f"+{len(b3_emails)} emails | +{len(b3_urls)} URLs")

        # ── Stage C: Email pivot & breach ──────────────────────────
        if "C" in active_stages and (emails or (usernames and domain)):
            self.log.info(f"[C] Email pivot — {len(STAGE_C_EMAIL_TOOLS)} tools")
            seed_email = sorted(emails)[0] if emails else f"{sorted(usernames)[0]}@{domain}" if domain else ""
            c_res = await self._run_stage("C", STAGE_C_EMAIL_TOOLS,
                {**_tokens(), "email": seed_email}, timeout, concurrency, profile)
            self._merge_into(c_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(c_res["tool_runs"]); events.extend(c_res["events"])
            stage_results["C"] = c_res["summary"]

        # ── Stage D: Social media ──────────────────────────────────
        if "D" in active_stages and usernames:
            self.log.info(f"[D] Social media — {len(STAGE_D_SOCIAL_TOOLS)} tools")
            d_res = await self._run_stage("D", STAGE_D_SOCIAL_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(d_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(d_res["tool_runs"]); events.extend(d_res["events"])
            stage_results["D"] = d_res["summary"]

        # ── Stage E: Domain intel ─────────────────────────────────
        if "E" in active_stages and domain:
            self.log.info(f"[E] Domain intel — {len(STAGE_E_DOMAIN_TOOLS)} tools")
            e_res = await self._run_stage("E", STAGE_E_DOMAIN_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(e_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(e_res["tool_runs"]); events.extend(e_res["events"])
            stage_results["E"] = e_res["summary"]

        # ── Stage F: Code/secrets ─────────────────────────────────
        if "F" in active_stages and (usernames or profile == "aggressive"):
            self.log.info(f"[F] Code/secrets — {len(STAGE_F_CODE_TOOLS)} tools")
            f_res = await self._run_stage("F", STAGE_F_CODE_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(f_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(f_res["tool_runs"]); events.extend(f_res["events"])
            stage_results["F"] = f_res["summary"]

        # ── Stage G: Phone OSINT ──────────────────────────────────
        if "G" in active_stages and (phones or (emails and profile in ("full","aggressive"))):
            self.log.info(f"[G] Phone OSINT — {len(STAGE_G_PHONE_TOOLS)} tools")
            g_res = await self._run_stage("G", STAGE_G_PHONE_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(g_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(g_res["tool_runs"]); events.extend(g_res["events"])
            stage_results["G"] = g_res["summary"]

        # ── Stage H: Dark web/breach ─────────────────────────────
        if "H" in active_stages and profile in ("full","aggressive"):
            self.log.info(f"[H] Dark web/breach — {len(STAGE_H_DARKWEB_TOOLS)} tools")
            h_res = await self._run_stage("H", STAGE_H_DARKWEB_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(h_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(h_res["tool_runs"]); events.extend(h_res["events"])
            stage_results["H"] = h_res["summary"]

        # ── Stage I: Geo/metadata ─────────────────────────────────
        if "I" in active_stages:
            self.log.info(f"[I] Geo/metadata — {len(STAGE_I_GEO_TOOLS)} tools")
            i_res = await self._run_stage("I", STAGE_I_GEO_TOOLS, _tokens(), timeout, concurrency, profile)
            self._merge_into(i_res, usernames, emails, names, urls, phones, max_pivots)
            tool_runs.extend(i_res["tool_runs"]); events.extend(i_res["events"])
            stage_results["I"] = i_res["summary"]

        # ── Stage J: Claude AI Correlation ───────────────────────
        ai_analysis: Dict[str, Any] = {}
        if "J" in active_stages and self.get("use_ai"):
            self.log.info("[J] Claude AI correlation analysis")
            ai_analysis = await self._claude_ai_stage(
                usernames=sorted(usernames), emails=sorted(emails),
                names=sorted(names), urls=sorted(urls)[:50],
                phones=sorted(phones), domain=domain,
                scraped_profiles=scraped_profiles[:10],
                tool_runs=tool_runs, stage_results=stage_results,
            )

        # ── Stage K: Report ──────────────────────────────────────
        correlation = self._build_correlation(usernames, emails, names, urls, tool_runs)

        payload = {
            "investigation": {
                "seed": {"username": self.get("username"), "email": self.get("email"),
                         "full_name": self.get("full_name"), "phone": self.get("phone"), "domain": domain},
                "profile": profile, "stages_executed": sorted(active_stages),
                "scraping_enabled": do_scrape,
                "generated_at": start_ts.isoformat(),
                "duration_seconds": round((datetime.utcnow()-start_ts).total_seconds(), 2),
            },
            "summary": {
                "total_tools_planned": sum(len(t) for t in [
                    STAGE_B_USERNAME_TOOLS, STAGE_C_EMAIL_TOOLS, STAGE_D_SOCIAL_TOOLS,
                    STAGE_E_DOMAIN_TOOLS, STAGE_F_CODE_TOOLS, STAGE_G_PHONE_TOOLS,
                    STAGE_H_DARKWEB_TOOLS, STAGE_I_GEO_TOOLS]),
                "total_tools_executed":  len([t for t in tool_runs if t.get("status") != "missing"]),
                "total_tools_missing":   len([t for t in tool_runs if t.get("status") == "missing"]),
                "profiles_scraped": len(scraped_profiles),
                "profiles_succeeded": sum(1 for p in scraped_profiles if p.get("scrape_success")),
                "unique_usernames": len(usernames),
                "unique_emails": len(emails),
                "unique_names": len(names),
                "unique_urls": len(urls),
                "unique_phones": len(phones),
            },
            "entities": {
                "usernames": sorted(usernames), "emails": sorted(emails),
                "names": sorted(names), "urls": sorted(urls), "phones": sorted(phones),
            },
            "scraped_profiles": scraped_profiles,
            "stage_results": stage_results,
            "correlation": correlation,
            "ai_analysis": ai_analysis,
            "tool_runs": tool_runs,
            "tool_universe": {
                "stage_b_username":  [t[0] for t in STAGE_B_USERNAME_TOOLS],
                "stage_c_email":     [t[0] for t in STAGE_C_EMAIL_TOOLS],
                "stage_d_social":    [t[0] for t in STAGE_D_SOCIAL_TOOLS],
                "stage_e_domain":    [t[0] for t in STAGE_E_DOMAIN_TOOLS],
                "stage_f_code":      [t[0] for t in STAGE_F_CODE_TOOLS],
                "stage_g_phone":     [t[0] for t in STAGE_G_PHONE_TOOLS],
                "stage_h_darkweb":   [t[0] for t in STAGE_H_DARKWEB_TOOLS],
                "stage_i_geo":       [t[0] for t in STAGE_I_GEO_TOOLS],
                "stage_b2_scrapers": list(STAGE_B2_SCRAPER_COMMANDS.keys()),
                "additional_socmint_tools": ADDITIONAL_SOCMINT_TOOLS,
            },
        }

        out_file = self._export(payload)

        findings = []
        for prof in scraped_profiles:
            if prof.get("scrape_success"):
                findings.append(self.make_finding(
                    title=f"Profile scraped: {prof.get('username','')} on {prof.get('platform','')}",
                    target=self.get("username") or domain,
                    severity=Severity.INFO,
                    description=(
                        f"Bio: {prof.get('bio','')[:120]} | "
                        f"Followers: {prof.get('followers','?')} | "
                        f"Location: {prof.get('location','?')}"
                    ),
                    evidence={k: v for k, v in prof.items() if k != "raw_text"},
                    tags=["osint", "socmint", "scraping", prof.get("platform","unknown")],
                ))
        for email in sorted(emails):
            findings.append(self.make_finding(
                title=f"Email discovered: {email}",
                target=self.get("username") or domain,
                severity=Severity.LOW,
                description="Email associated with target identity",
                evidence={"email": email},
                tags=["osint", "email"],
            ))
        for url in sorted(urls)[:50]:
            findings.append(self.make_finding(
                title=f"Account/profile: {url[:80]}",
                target=self.get("username") or domain,
                severity=Severity.INFO,
                description="Online presence discovered",
                evidence={"url": url},
                tags=["osint", "identity"],
            ))

        return ModuleResult(
            success=True,
            output={
                "output_file": out_file,
                "summary": payload["summary"],
                "entities": payload["entities"],
                "scraped_profiles_count": len(scraped_profiles),
                "correlation": correlation,
            },
            findings=findings,
            raw_output=json.dumps({"summary": payload["summary"]}, indent=2),
        )

    # ─────────────────────────────────────────────────────────────────
    # Stage B2: Deep account scraping
    # ─────────────────────────────────────────────────────────────────

    async def _scrape_found_accounts(
        self,
        account_urls: List[str],
        seed_username: str,
        timeout: int,
        concurrency: int,
        profile: str,
    ) -> Tuple[List[Dict[str, Any]], Dict[str, List[str]]]:
        """
        Scrape every discovered account URL.
        Returns (list_of_profile_dicts, new_entities_dict).
        """
        sem = asyncio.Semaphore(concurrency)
        profiles: List[ScrapedProfile] = []
        new_entities: Dict[str, List[str]] = {"emails": [], "urls": [], "usernames": [], "phones": []}

        async def _scrape_one(url: str) -> ScrapedProfile:
            async with sem:
                platform, stype = _detect_platform(url)
                username_on_platform = _extract_username_from_url(url, platform)

                # 1. Try platform-specific CLI tool first
                if stype != "http" and stype in STAGE_B2_SCRAPER_COMMANDS:
                    cli_profile = await self._scrape_with_cli(
                        url=url, platform=platform, stype=stype,
                        username=username_on_platform or seed_username,
                        timeout=timeout,
                    )
                    if cli_profile.scrape_success:
                        return cli_profile

                # 2. Try JSON/API endpoint for known platforms
                api_profile = await self._scrape_platform_api(
                    url=url, platform=platform,
                    username=username_on_platform or seed_username,
                    timeout=timeout,
                )
                if api_profile.scrape_success:
                    return api_profile

                # 3. Generic HTTP scrape fallback
                return await self._scrape_profile_http(
                    url=url, platform=platform,
                    username=username_on_platform,
                    timeout=timeout,
                )

        tasks = [_scrape_one(u) for u in account_urls]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        for r in results:
            if isinstance(r, ScrapedProfile):
                profiles.append(r)
                if r.scrape_success:
                    new_entities["emails"].extend(r.emails_found)
                    new_entities["urls"].extend(r.linked_urls)
                    new_entities["usernames"].extend(r.linked_accounts)
                    new_entities["phones"].extend(r.phones_found)
            # Silently skip exceptions — graceful degradation

        return [p.to_dict() for p in profiles], new_entities

    async def _scrape_with_cli(
        self,
        url: str, platform: str, stype: str, username: str, timeout: int,
    ) -> ScrapedProfile:
        """Run platform-specific CLI scraper and parse output."""
        prof = ScrapedProfile(url=url, platform=platform, username=username, scrape_method=f"cli:{stype}")
        if not username:
            return prof
        cmd_template = STAGE_B2_SCRAPER_COMMANDS.get(stype, [])
        if not cmd_template:
            return prof
        binary = cmd_template[0]
        if not shutil.which(binary):
            prof.error = f"binary not found: {binary}"
            return prof
        cmd = [t.replace("{username}", username).replace("{url}", url) for t in cmd_template]
        try:
            stdout, stderr, rc = await self.run_command_async(cmd, timeout=timeout)
            if stdout:
                prof = self._parse_cli_output(prof, stdout, platform)
                prof.scrape_success = True
        except Exception as exc:
            prof.error = str(exc)[:200]
        return prof

    def _parse_cli_output(self, prof: ScrapedProfile, text: str, platform: str) -> ScrapedProfile:
        """Parse CLI tool output to populate a ScrapedProfile."""
        prof.raw_text = text[:3000]
        prof.emails_found  = list(dict.fromkeys(EMAIL_RE.findall(text)))[:20]
        prof.linked_urls   = list(dict.fromkeys(URL_RE.findall(text)))[:30]

        patterns = _HTML_PATTERNS.get(platform, _HTML_PATTERNS["default"])
        for key, pat in patterns:
            m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
            if m:
                val = m.group(1).strip()
                if key == "bio" or key == "bio2":
                    if not prof.bio:
                        prof.bio = val
                elif key == "name":
                    if not prof.display_name:
                        prof.display_name = val
                elif key in ("followers", "following", "posts", "repos"):
                    try:
                        setattr(prof, key if key != "repos" else "posts", int(val.replace(",","").replace(".","")[:10]))
                    except Exception:
                        pass
                elif key == "location":
                    prof.location = val
                elif key == "website":
                    prof.website = val
                elif key == "joined":
                    prof.joined = val
        return prof

    async def _scrape_platform_api(
        self,
        url: str, platform: str, username: str, timeout: int,
    ) -> ScrapedProfile:
        """Hit unofficial JSON/API endpoints for known platforms."""
        prof = ScrapedProfile(url=url, platform=platform, username=username, scrape_method="api")
        if not username:
            return prof

        # Platform-specific API endpoints
        api_urls: List[str] = []
        if platform == "github":
            api_urls = [f"https://api.github.com/users/{username}"]
        elif platform == "reddit":
            api_urls = [f"https://www.reddit.com/user/{username}/about.json"]
        elif platform in ("twitter", "x"):
            # Public API v2 — no auth required for basic lookup attempt
            api_urls = [f"https://cdn.syndication.twimg.com/widgets/followbutton/info.json?screen_names={username}"]
        elif platform == "hackernews":
            api_urls = [f"https://hacker-news.firebaseio.com/v0/user/{username}.json"]
        elif platform == "mastodon":
            # Try to parse instance from URL
            parsed = urlparse(url)
            if parsed.netloc:
                api_urls = [f"https://{parsed.netloc}/api/v1/accounts/lookup?acct={username}"]

        if not api_urls:
            return prof

        try:
            import httpx
        except ImportError:
            return prof

        for api_url in api_urls:
            try:
                async with httpx.AsyncClient(timeout=timeout, follow_redirects=True,
                                              headers={"User-Agent": "Mozilla/5.0"}) as client:
                    resp = await client.get(api_url)
                    if resp.status_code == 200:
                        try:
                            data = resp.json()
                            prof = self._parse_json_profile(prof, data, platform)
                            prof.scrape_success = True
                            return prof
                        except Exception:
                            pass
            except Exception:
                pass
        return prof

    def _parse_json_profile(self, prof: ScrapedProfile, data: Any, platform: str) -> ScrapedProfile:
        """Parse JSON API response into ScrapedProfile fields."""
        if not isinstance(data, dict):
            if isinstance(data, list) and data:
                data = data[0]
            else:
                return prof

        # GitHub
        if platform == "github":
            prof.display_name = data.get("name", "")
            prof.bio          = data.get("bio", "") or ""
            prof.location     = data.get("location", "") or ""
            prof.website      = data.get("blog", "") or ""
            prof.followers    = data.get("followers")
            prof.following    = data.get("following")
            prof.posts        = data.get("public_repos")
            prof.joined       = data.get("created_at", "")
            email             = data.get("email", "")
            if email:
                prof.emails_found.append(email)
            prof.linked_urls.extend(
                [u for u in [data.get("html_url",""), data.get("blog","")] if u]
            )

        # Reddit
        elif platform == "reddit":
            d = data.get("data", data)
            prof.display_name = d.get("name", "")
            prof.bio          = d.get("subreddit", {}).get("public_description", "") or d.get("subreddit",{}).get("description","") or ""
            prof.posts        = d.get("link_karma", 0)
            # Convert timestamp
            created = d.get("created_utc") or d.get("created")
            if created:
                try:
                    prof.joined = datetime.utcfromtimestamp(float(created)).strftime("%Y-%m-%d")
                except Exception:
                    pass
            prof.verified     = d.get("verified", False)

        # Hacker News
        elif platform == "hackernews":
            prof.display_name = data.get("id", "")
            prof.bio          = data.get("about", "") or ""
            prof.joined       = str(data.get("created", ""))
            prof.posts        = data.get("submitted", 0) if isinstance(data.get("submitted"), int) else len(data.get("submitted", []))

        # Twitter syndication
        elif platform in ("twitter","x"):
            if isinstance(data, list) and data:
                d = data[0]
                prof.followers    = d.get("followers_count")
                prof.display_name = d.get("name", "")

        # Mastodon
        elif platform == "mastodon":
            prof.display_name = data.get("display_name", "")
            prof.bio          = re.sub(r"<[^>]+>", " ", data.get("note", "")).strip()
            prof.followers    = data.get("followers_count")
            prof.following    = data.get("following_count")
            prof.posts        = data.get("statuses_count")
            prof.website      = data.get("url", "")

        # Generic JSON key scan
        else:
            for key in ("bio","description","about","summary","blurb"):
                if data.get(key):
                    prof.bio = str(data[key])[:400]; break
            for key in ("name","display_name","full_name","realname","displayName"):
                if data.get(key):
                    prof.display_name = str(data[key]); break
            for key in ("followers","followers_count","follower_count"):
                if isinstance(data.get(key), (int,float)):
                    prof.followers = int(data[key]); break
            for key in ("location","city","country"):
                if data.get(key):
                    prof.location = str(data[key]); break
            for key in ("website","url","homepage","blog","link"):
                if data.get(key) and str(data[key]).startswith("http"):
                    prof.website = str(data[key]); break
            email = data.get("email","")
            if email and "@" in str(email):
                prof.emails_found.append(str(email))

        # Always scan raw JSON for emails
        raw = json.dumps(data)
        prof.emails_found = list(dict.fromkeys(prof.emails_found + EMAIL_RE.findall(raw)))[:20]
        prof.linked_urls  = list(dict.fromkeys(prof.linked_urls  + [u for u in URL_RE.findall(raw) if "api." not in u]))[:20]
        return prof

    async def _scrape_profile_http(
        self,
        url: str, platform: str, username: str, timeout: int,
    ) -> ScrapedProfile:
        """Generic async HTTP scraper — fetches the profile URL and extracts structured data."""
        prof = ScrapedProfile(url=url, platform=platform, username=username, scrape_method="http")
        try:
            import httpx
        except ImportError:
            prof.error = "httpx not installed"
            return prof

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
        }

        try:
            async with httpx.AsyncClient(
                timeout=timeout,
                follow_redirects=True,
                headers=headers,
                verify=False,
            ) as client:
                resp = await client.get(url)
                if resp.status_code not in (200, 203):
                    prof.error = f"HTTP {resp.status_code}"
                    return prof

                html = resp.text[:80000]  # limit to 80 KB
                prof.raw_text = html[:3000]

                # Extract emails & URLs from raw HTML
                prof.emails_found = list(dict.fromkeys(
                    e for e in EMAIL_RE.findall(html)
                    if not e.endswith((".png",".jpg",".gif",".svg"))
                ))[:20]
                all_urls = URL_RE.findall(html)
                prof.linked_urls = list(dict.fromkeys(
                    u for u in all_urls
                    if any(s in u for s in ["twitter","instagram","github","linkedin","youtube","tiktok","facebook"])
                    or (platform == "unknown")
                ))[:30]

                # Phone numbers (loose heuristic)
                raw_phones = PHONE_RE.findall(html)
                prof.phones_found = list(dict.fromkeys(
                    p.strip() for p in raw_phones
                    if len(re.sub(r"\D","",p)) >= 8
                ))[:5]

                # Apply per-platform HTML patterns
                patterns = _HTML_PATTERNS.get(platform, _HTML_PATTERNS["default"])
                for key, pat in patterns:
                    m = re.search(pat, html, re.IGNORECASE | re.DOTALL)
                    if m:
                        val = m.group(1).strip()
                        if key in ("bio","bio2") and not prof.bio:
                            prof.bio = re.sub(r"<[^>]+>", " ", val).strip()[:400]
                        elif key == "name" and not prof.display_name:
                            prof.display_name = re.sub(r"<[^>]+>", " ", val).strip()[:100]
                        elif key == "followers":
                            try:
                                prof.followers = int(val.replace(",","").replace(".","")[:10])
                            except Exception:
                                pass
                        elif key == "following":
                            try:
                                prof.following = int(val.replace(",","").replace(".","")[:10])
                            except Exception:
                                pass
                        elif key == "posts":
                            try:
                                prof.posts = int(val.replace(",","").replace(".","")[:10])
                            except Exception:
                                pass
                        elif key == "location" and not prof.location:
                            prof.location = val[:100]
                        elif key == "website" and not prof.website:
                            prof.website = val[:200]
                        elif key == "joined" and not prof.joined:
                            prof.joined = val[:50]

                # Detect linked social accounts from href attributes
                linked = re.findall(
                    r'href=["\'](https?://(?:twitter|x|instagram|github|linkedin|youtube|tiktok|facebook|twitch)\.(?:com|tv)/[^"\'?\s]+)',
                    html
                )
                prof.linked_accounts = list(dict.fromkeys(
                    _extract_username_from_url(u, _detect_platform(u)[0])
                    for u in linked if _extract_username_from_url(u, _detect_platform(u)[0])
                ))[:15]

                # Try to grab a few post snippets (look for <p> tags with real content)
                post_matches = re.findall(r"<p[^>]*>([^<]{40,300})</p>", html)
                prof.post_samples = [re.sub(r"\s+"," ",p).strip() for p in post_matches[:5]]

                prof.scrape_success = True

        except Exception as exc:
            prof.error = str(exc)[:200]

        return prof

    # ─────────────────────────────────────────────────────────────────
    # Stage B3: Multi-engine web search
    # ─────────────────────────────────────────────────────────────────

    async def _run_web_search_stage(
        self,
        queries: List[str],
        engines: str,
        results_per_engine: int,
        fetch_pages: bool,
        max_fetch: int,
        timeout: int,
        delay: float,
        google_key: str = "",
        google_cx: str = "",
        brave_key: str = "",
    ) -> Tuple[List[Dict], Set[str], Set[str], Set[str]]:
        """Run multi-engine web search for all queries and collect entities."""
        from framework.modules.osint.web_search_scraper import (
            _scrape_duckduckgo, _scrape_bing, _scrape_brave,
            _scrape_google, _scrape_yahoo, _scrape_startpage,
            _fetch_result_page, SearchResult,
        )

        engine_list = [e.strip().lower() for e in engines.split(",") if e.strip()]
        engine_funcs = {
            "duckduckgo": lambda q: _scrape_duckduckgo(q, results_per_engine, timeout),
            "bing":        lambda q: _scrape_bing(q, results_per_engine, timeout),
            "brave":       lambda q: _scrape_brave(q, results_per_engine, timeout, brave_key),
            "google":      lambda q: _scrape_google(q, results_per_engine, timeout, google_key, google_cx),
            "yahoo":       lambda q: _scrape_yahoo(q, results_per_engine, timeout),
            "startpage":   lambda q: _scrape_startpage(q, results_per_engine, timeout),
        }

        all_results: List[SearchResult] = []
        seen_urls: Set[str] = set()

        for query in queries:
            for engine in engine_list:
                fn = engine_funcs.get(engine)
                if not fn:
                    continue
                try:
                    results = await fn(query)
                    for r in results:
                        norm = r.url.rstrip("/").lower()
                        if norm not in seen_urls:
                            seen_urls.add(norm)
                            all_results.append(r)
                except Exception as exc:
                    self.log.debug(f"  web search {engine} error: {exc}")
                if delay > 0:
                    await asyncio.sleep(delay)

        # Optionally fetch result pages
        if fetch_pages and all_results:
            sem = asyncio.Semaphore(4)
            to_fetch = [r for r in all_results if not r.page_fetched][:max_fetch]
            async def _guarded(r: SearchResult) -> SearchResult:
                async with sem:
                    return await _fetch_result_page(r, timeout)
            fetched = await asyncio.gather(*[_guarded(r) for r in to_fetch], return_exceptions=True)
            for i, upd in enumerate(fetched):
                if isinstance(upd, SearchResult):
                    all_results[all_results.index(to_fetch[i])] = upd

        # Aggregate extracted entities
        new_emails: Set[str] = set()
        new_urls:   Set[str] = set()
        new_social: Set[str] = set()
        for r in all_results:
            new_emails.update(r.page_emails)
            new_urls.add(r.url)
            new_social.update(r.page_social)
            # Also scan snippet and title for emails
            for text in (r.snippet, r.title, r.page_meta_desc):
                new_emails.update(EMAIL_RE.findall(text))

        return [r.to_dict() for r in all_results], new_emails, new_urls, new_social

    # ─────────────────────────────────────────────────────────────────
    # General stage runner
    # ─────────────────────────────────────────────────────────────────

    async def _run_stage(
        self,
        stage: str,
        tool_specs: List[Tuple],
        tokens: Dict[str, str],
        timeout: int,
        concurrency: int,
        profile: str,
    ) -> Dict[str, Any]:
        sem = asyncio.Semaphore(concurrency)
        tool_runs: List[Dict] = []
        events:    List[Dict] = []

        if profile == "core":
            specs = tool_specs[:5]
        elif profile == "full":
            specs = tool_specs[:10]
        else:
            specs = tool_specs

        async def _run_one(spec: Tuple) -> None:
            tool_name, binary, cmd_template, description = spec
            async with sem:
                cmd = self._render_cmd(cmd_template, tokens)
                if not cmd:
                    tool_runs.append({"tool": tool_name, "stage": stage, "status": "skipped-empty-cmd"})
                    return
                if not shutil.which(binary):
                    tool_runs.append({"tool": tool_name, "stage": stage, "status": "missing",
                                      "binary": binary, "description": description})
                    return
                try:
                    stdout, stderr, rc = await self.run_command_async(cmd, timeout=timeout)
                    parsed = self._parse_output(stdout + "\n" + stderr)
                    status = "ok" if rc == 0 else "error"
                    tool_runs.append({
                        "tool": tool_name, "stage": stage, "status": status, "rc": rc,
                        "emails_found": len(parsed["emails"]),
                        "urls_found": len(parsed["urls"]),
                        "usernames_found": len(parsed["usernames"]),
                    })
                    events.append({"tool": tool_name, "stage": stage,
                                   "parsed": parsed, "stdout_preview": stdout[:800]})
                except asyncio.TimeoutError:
                    tool_runs.append({"tool": tool_name, "stage": stage, "status": "timeout"})
                except Exception as exc:
                    tool_runs.append({"tool": tool_name, "stage": stage, "status": "exception",
                                      "error": str(exc)[:200]})

        await asyncio.gather(*[_run_one(s) for s in specs], return_exceptions=True)

        ok       = len([t for t in tool_runs if t["stage"]==stage and t["status"]=="ok"])
        executed = len([t for t in tool_runs if t["stage"]==stage and t["status"] not in ("missing","skipped-empty-cmd")])
        return {
            "tool_runs": tool_runs, "events": events,
            "summary": {"stage": stage, "tools_executed": executed, "tools_ok": ok},
        }

    def _render_cmd(self, template: List[str], tokens: Dict[str, str]) -> List[str]:
        rendered = []
        for tok in template:
            for k, v in tokens.items():
                tok = tok.replace(f"{{{k}}}", v or "")
            rendered.append(tok)
        rendered = [t for t in rendered if t]
        if any(re.search(r"\{[a-z_]+\}", t) for t in rendered):
            return []
        return rendered

    def _parse_output(self, text: str) -> Dict[str, List[str]]:
        emails    = list(dict.fromkeys(EMAIL_RE.findall(text)))
        urls      = list(dict.fromkeys(URL_RE.findall(text)))
        usernames: Set[str] = set()
        names:     Set[str] = set()
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            lower = line.lower()
            if any(m in lower for m in ("[+]","found:","exists:","valid:","registered:","FOUND")):
                for tok in USERNAME_RE.findall(line):
                    if self._is_username(tok):
                        usernames.add(tok)
            if any(m in lower for m in ("name:","real name:","full name:")):
                for tok in re.findall(r"[A-Z][a-z]+ [A-Z][a-z]+", line):
                    names.add(tok)
        return {
            "emails": emails[:100], "urls": urls[:200],
            "usernames": list(usernames)[:50], "names": list(names)[:30],
        }

    def _is_username(self, token: str) -> bool:
        return (
            3 <= len(token) <= 30
            and not token.startswith("http")
            and "@" not in token
            and not token.isdigit()
            and token.lower() not in {
                "true","false","none","null","error","info","found","valid",
                "http","https","json","html","text","type","user","name",
            }
        )

    def _merge_into(
        self,
        stage_res: Dict,
        usernames: Set[str], emails: Set[str], names: Set[str],
        urls: Set[str], phones: Set[str], max_pivots: int,
    ) -> None:
        for event in stage_res.get("events", []):
            parsed = event.get("parsed", {})
            usernames.update(list(parsed.get("usernames", []))[:max_pivots])
            emails.update(   list(parsed.get("emails",    []))[:max_pivots])
            names.update(    list(parsed.get("names",     []))[:max_pivots])
            urls.update(     list(parsed.get("urls",      []))[:max_pivots*4])

    # ─────────────────────────────────────────────────────────────────
    # Stage J: Claude AI Correlation
    # ─────────────────────────────────────────────────────────────────

    async def _claude_ai_stage(
        self,
        usernames: List[str], emails: List[str], names: List[str],
        urls: List[str], phones: List[str], domain: str,
        scraped_profiles: List[Dict],
        tool_runs: List[Dict], stage_results: Dict,
    ) -> Dict[str, Any]:
        try:
            import httpx
        except ImportError:
            return self._rule_based_correlation(usernames, emails, names, urls)

        # Summarise scraped profiles for the AI prompt
        profile_summaries = []
        for p in scraped_profiles[:8]:
            if p.get("scrape_success"):
                profile_summaries.append({
                    "platform":      p.get("platform"),
                    "username":      p.get("username"),
                    "display_name":  p.get("display_name"),
                    "bio":           p.get("bio","")[:200],
                    "followers":     p.get("followers"),
                    "location":      p.get("location"),
                    "website":       p.get("website"),
                    "joined":        p.get("joined"),
                    "emails_found":  p.get("emails_found",[])[:5],
                    "linked_accounts": p.get("linked_accounts",[])[:5],
                })

        investigation_data = {
            "seed_username":      self.get("username"),
            "seed_email":         self.get("email"),
            "seed_full_name":     self.get("full_name"),
            "seed_domain":        domain,
            "discovered_usernames": usernames[:20],
            "discovered_emails":    emails[:15],
            "discovered_names":     names[:10],
            "sample_urls":          urls[:20],
            "discovered_phones":    phones[:5],
            "scraped_profiles":     profile_summaries,
            "stages_completed":     list(stage_results.keys()),
            "tools_executed":       len([t for t in tool_runs if t.get("status") == "ok"]),
            "tools_missing":        len([t for t in tool_runs if t.get("status") == "missing"]),
        }

        prompt = f"""You are an expert OSINT analyst performing an identity correlation.

Investigation data (Stage B–I results + scraped profile data):
{json.dumps(investigation_data, indent=2)}

Produce a structured correlation analysis in JSON with these keys:
- confidence_score (0-100 integer): How confident this is one individual
- identity_summary: Most likely real identity and key facts
- primary_aliases: The most-used usernames/handles
- cross_platform_timeline: Chronological account creation pattern
- location_indicators: Geographic clues across all sources
- bio_consistency_score (0-100): How consistent the bios are across platforms
- anomalies: Suspicious discrepancies (e.g. different names, location jumps)
- risk_level: LOW / MEDIUM / HIGH / CRITICAL (exposure level)
- top_pivots: Top 5 recommended next investigative actions
- linked_identity_graph: Object mapping each username to its confirmed platforms
- ai_model_used: Model identifier

Return only valid JSON, no markdown.
"""

        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"Content-Type": "application/json", "anthropic-version": "2023-06-01"},
                    json={
                        "model": "claude-sonnet-4-20250514",
                        "max_tokens": 2048,
                        "messages": [{"role": "user", "content": prompt}],
                    },
                )
                if response.status_code == 200:
                    text = response.json()["content"][0]["text"]
                    try:
                        result = json.loads(text)
                        result["analysis_type"] = "claude_ai"
                        result["ai_model_used"]  = "claude-sonnet-4-20250514"
                        return result
                    except json.JSONDecodeError:
                        m = re.search(r"\{.*\}", text, re.DOTALL)
                        if m:
                            try:
                                r = json.loads(m.group())
                                r["analysis_type"] = "claude_ai"
                                return r
                            except Exception:
                                pass
                        return {"analysis_type": "claude_ai_raw", "raw": text[:2000]}
        except Exception as e:
            self.log.warning(f"Claude AI stage error: {e}")

        return self._rule_based_correlation(usernames, emails, names, urls)

    def _rule_based_correlation(self, usernames: List[str], emails: List[str], names: List[str], urls: List[str]) -> Dict[str, Any]:
        confidence = min(100,
            (20 if len(urls)  > 10 else len(urls) * 2) +
            (15 if emails     else 0) +
            (10 if len(usernames) > 1 else 0) +
            (10 if names      else 0)
        )
        connections = []
        seed_un = (self.get("username") or "").lower()
        for email in emails:
            local = email.split("@")[0].lower()
            if seed_un and (seed_un in local or local in seed_un):
                confidence = min(100, confidence + 20)
                connections.append(f"Email {email} correlates with username {seed_un}")
        return {
            "analysis_type": "rule_based",
            "confidence_score": confidence,
            "cross_platform_connections": connections,
            "unique_profiles": len(urls),
            "risk_level": "HIGH" if len(urls) > 20 else "MEDIUM" if len(urls) > 5 else "LOW",
            "top_pivots": [
                "Review discovered emails for breach exposure",
                "Cross-reference usernames across remaining platforms",
                "Investigate bio/location consistency across scraped profiles",
                "Check common passwords in credential dumps",
                "Map social graph between discovered accounts",
            ],
        }

    def _build_correlation(self, usernames: Set, emails: Set, names: Set, urls: Set, tool_runs: List) -> Dict:
        ok       = len([t for t in tool_runs if t.get("status") == "ok"])
        executed = max(1, len([t for t in tool_runs if t.get("status") not in ("missing",)]))
        return {
            "confidence_score": min(100, int((ok/executed)*40 + min(len(urls),60)*0.5 + min(len(emails),20))),
            "deduplication": {"unique_usernames": len(usernames), "unique_emails": len(emails),
                              "unique_names": len(names), "unique_urls": len(urls)},
            "execution_health": {"tools_ok": ok, "tools_executed": executed,
                                 "success_ratio": round(ok/executed, 3)},
        }

    # ─────────────────────────────────────────────────────────────────
    # Stage K: Export
    # ─────────────────────────────────────────────────────────────────

    def _export(self, payload: Dict[str, Any]) -> str:
        fmt      = self.get("output_format")
        explicit = self.get("output_file")
        ts       = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        target   = Path(explicit) if explicit else Path("data") / f"identity_fusion_{ts}.{fmt}"
        target.parent.mkdir(parents=True, exist_ok=True)

        if fmt == "json":
            target.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
            return str(target)

        if fmt == "csv":
            rows = self._flatten_rows(payload)
            with target.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=["type","value","stage","platform"])
                writer.writeheader(); writer.writerows(rows)
            return str(target)

        if fmt == "xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                wb = Workbook()
                # Summary
                ws = wb.active; ws.title = "Summary"
                ws["A1"] = "RTF Identity Fusion v4.0 Report"
                ws["A1"].font = Font(bold=True, size=14, color="CC0000")
                r = 3
                for k, v in payload.get("summary",{}).items():
                    ws.cell(r,1,k).font = Font(bold=True); ws.cell(r,2,str(v)); r+=1
                # Entities
                we = wb.create_sheet("Entities"); we.append(["Type","Value"])
                for etype, vals in payload.get("entities",{}).items():
                    for val in vals:
                        we.append([etype, val])
                # Scraped profiles
                ws2 = wb.create_sheet("Scraped Profiles")
                cols = ["platform","username","display_name","bio","followers","following","posts",
                        "location","website","joined","verified","emails_found","linked_accounts","scrape_method","scrape_success"]
                ws2.append(cols)
                for p in payload.get("scraped_profiles",[]):
                    ws2.append([str(p.get(c,""))[:150] for c in cols])
                # Tool runs
                wt = wb.create_sheet("Tool Runs"); wt.append(["Tool","Stage","Status","Emails","URLs"])
                for t in payload.get("tool_runs",[]):
                    wt.append([t.get("tool",""),t.get("stage",""),t.get("status",""),
                                t.get("emails_found",""),t.get("urls_found","")])
                wb.save(str(target)); return str(target)
            except ImportError:
                pass
            fallback = target.with_suffix(".json")
            fallback.write_text(json.dumps(payload, indent=2, default=str)); return str(fallback)

        if fmt == "pdf":
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                from reportlab.lib import colors
                doc = SimpleDocTemplate(str(target), pagesize=letter)
                styles = getSampleStyleSheet(); story = []
                story.append(Paragraph("RTF Identity Fusion v4.0 Report", styles["Title"]))
                story.append(Spacer(1,12))
                for k,v in payload.get("summary",{}).items():
                    story.append(Paragraph(f"<b>{k}:</b> {v}", styles["Normal"]))
                story.append(Spacer(1,12))
                story.append(Paragraph("Scraped Profiles", styles["Heading2"]))
                for p in payload.get("scraped_profiles",[]):
                    if p.get("scrape_success"):
                        story.append(Paragraph(
                            f"<b>{p.get('platform','?').upper()}</b> — {p.get('username','')} | "
                            f"Followers: {p.get('followers','?')} | Location: {p.get('location','?')}",
                            styles["Normal"]
                        ))
                        if p.get("bio"):
                            story.append(Paragraph(f"  Bio: {p.get('bio','')[:200]}", styles["Italic"]))
                        story.append(Spacer(1,6))
                doc.build(story); return str(target)
            except ImportError:
                pass
            fallback = target.with_suffix(".html")
            fallback.write_text(self._build_html(payload)); return str(fallback)

        if fmt == "html":
            target.write_text(self._build_html(payload), encoding="utf-8"); return str(target)

        fallback = target.with_suffix(".json")
        fallback.write_text(json.dumps(payload, indent=2, default=str)); return str(fallback)

    def _flatten_rows(self, payload: Dict) -> List[Dict]:
        rows: List[Dict] = []
        for etype, vals in payload.get("entities",{}).items():
            for v in vals:
                rows.append({"type": etype, "value": str(v), "stage": "merged", "platform": ""})
        for p in payload.get("scraped_profiles",[]):
            if p.get("scrape_success"):
                rows.append({"type":"scraped_bio","value":p.get("bio","")[:200],"stage":"B2","platform":p.get("platform","")})
                for e in p.get("emails_found",[]):
                    rows.append({"type":"scraped_email","value":e,"stage":"B2","platform":p.get("platform","")})
                for a in p.get("linked_accounts",[]):
                    rows.append({"type":"linked_account","value":a,"stage":"B2","platform":p.get("platform","")})
        return rows

    def _build_html(self, payload: Dict) -> str:
        summary     = payload.get("summary", {})
        entities    = payload.get("entities", {})
        ai          = payload.get("ai_analysis", {})
        investigation = payload.get("investigation", {})
        seed        = investigation.get("seed", {})
        profiles    = payload.get("scraped_profiles", [])
        scraped_ok  = [p for p in profiles if p.get("scrape_success")]

        # Build scraped profiles section
        profiles_html = ""
        for p in scraped_ok:
            platform = p.get("platform","unknown").upper()
            color_map = {
                "TWITTER":"#1DA1F2","INSTAGRAM":"#E1306C","GITHUB":"#333","REDDIT":"#FF4500",
                "LINKEDIN":"#0077B5","YOUTUBE":"#FF0000","TIKTOK":"#010101","FACEBOOK":"#1877F2",
                "GITHUB":"#333","MASTODON":"#6364FF","DEFAULT":"#888",
            }
            clr = color_map.get(platform, color_map["DEFAULT"])
            emails_str  = ", ".join(p.get("emails_found",[])[:5]) or "—"
            linked_str  = ", ".join(p.get("linked_accounts",[])[:5]) or "—"
            phones_str  = ", ".join(p.get("phones_found",[])[:3]) or "—"
            post_str    = "<br>".join(f"<i>{s[:100]}</i>" for s in p.get("post_samples",[]))

            profiles_html += f"""
<div style="border:1px solid #333;border-left:5px solid {clr};background:#111;margin:10px 0;padding:14px;border-radius:6px;">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
    <span style="color:{clr};font-weight:bold;font-size:14px;">{platform}</span>
    <span style="color:#4cc9f0;font-size:13px;">@{p.get('username','')}</span>
    {'<span style="color:#f0c040;font-size:11px;">★ VERIFIED</span>' if p.get('verified') else ''}
  </div>
  <div style="color:#e0e0e0;font-size:13px;margin-bottom:6px;"><b>{p.get('display_name','')}</b></div>
  <div style="color:#aaa;font-size:12px;margin-bottom:8px;font-style:italic;">{p.get('bio','')[:300]}</div>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:8px;">
    <div style="background:#1a1a1a;padding:8px;border-radius:4px;text-align:center;">
      <div style="color:#dc2626;font-size:18px;font-weight:bold;">{p.get('followers','—'):,}</div>
      <div style="color:#666;font-size:10px;">Followers</div>
    </div>
    <div style="background:#1a1a1a;padding:8px;border-radius:4px;text-align:center;">
      <div style="color:#4cc9f0;font-size:18px;font-weight:bold;">{p.get('following','—'):,}</div>
      <div style="color:#666;font-size:10px;">Following</div>
    </div>
    <div style="background:#1a1a1a;padding:8px;border-radius:4px;text-align:center;">
      <div style="color:#a0c0ff;font-size:18px;font-weight:bold;">{p.get('posts','—'):,}</div>
      <div style="color:#666;font-size:10px;">Posts</div>
    </div>
    <div style="background:#1a1a1a;padding:8px;border-radius:4px;text-align:center;">
      <div style="color:#80ff80;font-size:12px;">{p.get('scrape_method','http')}</div>
      <div style="color:#666;font-size:10px;">Method</div>
    </div>
  </div>
  <table style="width:100%;font-size:11px;color:#888;">
    <tr><td><b>📍 Location</b></td><td>{p.get('location','—')}</td><td><b>🔗 Website</b></td><td>{p.get('website','—')[:50]}</td></tr>
    <tr><td><b>📅 Joined</b></td><td>{p.get('joined','—')}</td><td><b>📧 Emails found</b></td><td><code style="color:#f0c040">{emails_str}</code></td></tr>
    <tr><td><b>📱 Phones</b></td><td><code style="color:#80ff80">{phones_str}</code></td><td><b>🔗 Linked accounts</b></td><td><code style="color:#4cc9f0">{linked_str}</code></td></tr>
  </table>
  {f'<div style="margin-top:8px;font-size:11px;color:#666;border-top:1px solid #222;padding-top:6px;"><b>Post samples:</b><br>{post_str}</div>' if post_str else ''}
  <div style="margin-top:4px;"><a href="{p.get('url','')}" style="color:#4cc9f0;font-size:10px;">{p.get('url','')[:80]}</a></div>
</div>"""

        entity_html = ""
        for etype, vals in entities.items():
            if vals:
                items = "".join(f"<li><code style='color:#4cc9f0;font-size:11px;'>{v}</code></li>" for v in vals[:100])
                entity_html += f"<h3 style='color:#888;font-size:13px;'>{etype.capitalize()} ({len(vals)})</h3><ul>{items}</ul>"

        ai_html = ""
        if ai:
            ai_html = f"""
<h2 style='color:#dc2626;'>🤖 Claude AI Correlation Analysis</h2>
<div style='display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:12px 0;'>
  <div style='background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;'>
    <div style='font-size:32px;font-weight:bold;color:#dc2626;'>{ai.get('confidence_score','?')}</div>
    <div style='color:#888;font-size:11px;'>Confidence Score</div>
  </div>
  <div style='background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;'>
    <div style='font-size:18px;font-weight:bold;color:#{"dc2626" if ai.get("risk_level") in ("HIGH","CRITICAL") else "d97706" if ai.get("risk_level")=="MEDIUM" else "16a34a"};'>{ai.get('risk_level','?')}</div>
    <div style='color:#888;font-size:11px;'>Risk Level</div>
  </div>
  <div style='background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;'>
    <div style='font-size:18px;font-weight:bold;color:#4cc9f0;'>{ai.get('bio_consistency_score','?')}</div>
    <div style='color:#888;font-size:11px;'>Bio Consistency</div>
  </div>
</div>
<p style='color:#ccc;font-size:13px;'>{ai.get('identity_summary','')}</p>
<h3 style='color:#4cc9f0;'>Recommended Pivots</h3>
<ol style='color:#aaa;font-size:12px;'>{"".join(f"<li>{s}</li>" for s in (ai.get('top_pivots') or []))}</ol>
{f'<h3 style="color:#4cc9f0;">Anomalies</h3><p style="color:#f0c040;font-size:12px;">{ai.get("anomalies","—")}</p>' if ai.get("anomalies") else ""}
"""

        return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>RTF Identity Fusion v4.0 Report</title>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0a0a0a;color:#e0e0e0;margin:0;padding:0;}}
.header{{background:linear-gradient(135deg,#1a0505,#2d0000);padding:30px;border-bottom:3px solid #dc2626;}}
.container{{max-width:1200px;margin:0 auto;padding:24px;}}
ul{{column-count:2;column-gap:20px;list-style:none;padding:0;}}
h2{{color:#dc2626;border-bottom:1px solid #2a2a2a;padding-bottom:6px;margin-top:30px;}}
code{{background:#1a1a1a;padding:2px 6px;border-radius:3px;font-size:11px;}}
</style></head><body>
<div class="header">
  <h1 style="margin:0;color:#dc2626;font-size:26px;">⚔ RTF Identity Fusion v4.0 — Investigation Report</h1>
  <p style="color:#888;margin:6px 0 0;">Generated: {payload.get('investigation',{}).get('generated_at','?')} |
     Duration: {payload.get('investigation',{}).get('duration_seconds','?')}s |
     Profile: {investigation.get('profile','core').upper()} |
     Scraping: {'✓ ENABLED' if investigation.get('scraping_enabled') else '✗ DISABLED'}</p>
</div>
<div class="container">
  <h2>🎯 Seeds</h2>
  <table style="font-size:12px;"><tbody>
    {"".join(f"<tr><td style='color:#888;padding:3px 16px 3px 0;'><b>{k}</b></td><td><code>{v or 'N/A'}</code></td></tr>" for k,v in seed.items())}
  </tbody></table>

  <h2>📊 Summary</h2>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:12px 0;">
    <div style="background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;">
      <div style="font-size:28px;font-weight:bold;color:#dc2626;">{summary.get('profiles_scraped',0)}</div>
      <div style="color:#888;font-size:10px;">Profiles Scraped</div>
    </div>
    <div style="background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;">
      <div style="font-size:28px;font-weight:bold;color:#16a34a;">{summary.get('profiles_succeeded',0)}</div>
      <div style="color:#888;font-size:10px;">Scraped Successfully</div>
    </div>
    <div style="background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;">
      <div style="font-size:28px;font-weight:bold;color:#4cc9f0;">{summary.get('unique_emails',0)}</div>
      <div style="color:#888;font-size:10px;">Emails Discovered</div>
    </div>
    <div style="background:#111;border:1px solid #333;border-radius:6px;padding:12px;text-align:center;">
      <div style="font-size:28px;font-weight:bold;color:#f0c040;">{summary.get('unique_urls',0)}</div>
      <div style="color:#888;font-size:10px;">Profile URLs</div>
    </div>
  </div>

  {ai_html}

  <h2>🔍 Scraped Profiles ({len(scraped_ok)} succeeded)</h2>
  {profiles_html if profiles_html else '<p style="color:#555;">No profiles scraped yet. Run with scrape_accounts=true.</p>'}

  <h2>📋 All Discovered Entities</h2>
  {entity_html}

  <h2>🔧 Tool Execution ({len(payload.get('tool_runs',[]))} total)</h2>
  <div style="overflow-x:auto;">
  <table style="width:100%;font-size:11px;border-collapse:collapse;">
    <thead><tr style="border-bottom:1px solid #333;color:#666;">
      <th style="text-align:left;padding:6px;">Tool</th>
      <th style="padding:6px;">Stage</th><th style="padding:6px;">Status</th>
      <th style="padding:6px;">Emails</th><th style="padding:6px;">URLs</th>
    </tr></thead><tbody>
    {"".join(f"<tr style='border-bottom:1px solid #1a1a1a;'><td style='padding:5px 8px;color:#ccc;'>{t.get('tool','')}</td><td style='padding:5px 8px;text-align:center;color:#4cc9f0;'>{t.get('stage','')}</td><td style='padding:5px 8px;text-align:center;color:{'#16a34a' if t.get('status')=='ok' else '#dc2626' if t.get('status') in ('error','exception') else '#555'};'>{t.get('status','')}</td><td style='padding:5px 8px;text-align:center;'>{t.get('emails_found','')}</td><td style='padding:5px 8px;text-align:center;'>{t.get('urls_found','')}</td></tr>" for t in payload.get('tool_runs',[])[:120])}
    </tbody>
  </table>
  </div>
</div>
<footer><p style="color:#333;font-size:10px;text-align:center;margin-top:40px;padding:16px;">
RTF Identity Fusion v4.0 — For Authorized Investigations Only
</p></footer>
</body></html>"""
