"""
RedTeam Framework v2.0 - Professional Reporting Engine
Multi-format report generation: JSON, HTML, PDF, XLSX, Markdown.
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

from framework.core.logger import get_logger

log = get_logger("rtf.reporting")


class Severity(str, Enum):
    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


SEVERITY_ORDER = {
    Severity.CRITICAL: 0,
    Severity.HIGH: 1,
    Severity.MEDIUM: 2,
    Severity.LOW: 3,
    Severity.INFO: 4,
}

MITRE_TAGS = {
    "subdomain": "T1590.001 - Gather Victim Network Information",
    "port-scan": "T1046 - Network Service Discovery",
    "nuclei": "T1190 - Exploit Public-Facing Application",
    "sqli": "T1190 - Exploit Public-Facing Application",
    "xss": "T1059.007 - JavaScript Injection",
    "bloodhound": "T1482 - Domain Trust Discovery",
    "kerberoast": "T1558.003 - Kerberoasting",
    "asreproast": "T1558.004 - AS-REP Roasting",
    "aws": "T1526 - Cloud Service Discovery",
    "azure": "T1526 - Cloud Service Discovery",
    "credential": "T1078 - Valid Accounts",
    "password": "T1110 - Brute Force",
    "ldap": "T1018 - Remote System Discovery",
    "osint": "T1591 - Gather Victim Org Information",
    "email": "T1589.002 - Email Addresses",
    "username": "T1589.003 - Employee Names",
    "ssl": "T1590.003 - Network Trust Dependencies",
    "suid": "T1548.001 - Setuid and Setgid",
    "privesc": "T1068 - Exploitation for Privilege Escalation",
    "creds": "T1552 - Unsecured Credentials",
}


@dataclass
class Finding:
    title: str
    target: str
    severity: Severity = Severity.INFO
    description: str = ""
    category: str = "general"
    evidence: Dict[str, Any] = field(default_factory=dict)
    tags: List[str] = field(default_factory=list)
    cvss: Optional[float] = None
    cve: Optional[str] = None
    mitre: Optional[str] = None


class ReportEngine:
    """
    Generate professional engagement reports.
    Supports HTML, PDF, XLSX, Markdown, and JSON output formats.
    """

    def generate(
        self,
        title: str,
        findings: List[Finding],
        format: str = "html",
        output_path: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> str:
        metadata = metadata or {}
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = f"data/report_{ts}.{format}"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Enrich findings with MITRE tags
        for finding in findings:
            if not finding.mitre:
                for tag in finding.tags:
                    if tag.lower() in MITRE_TAGS:
                        finding.mitre = MITRE_TAGS[tag.lower()]
                        break

        # Sort by severity
        sorted_findings = sorted(findings, key=lambda f: SEVERITY_ORDER.get(f.severity, 99))

        if format == "html":
            return self._gen_html(title, sorted_findings, metadata, output_path)
        elif format == "pdf":
            return self._gen_pdf(title, sorted_findings, metadata, output_path)
        elif format == "xlsx":
            return self._gen_xlsx(title, sorted_findings, metadata, output_path)
        elif format in ("md", "markdown"):
            return self._gen_markdown(title, sorted_findings, metadata, output_path)
        elif format == "json":
            return self._gen_json(title, sorted_findings, metadata, output_path)
        else:
            return self._gen_html(title, sorted_findings, metadata, output_path)

    def _severity_counts(self, findings: List[Finding]) -> Dict[str, int]:
        counts = {s.value: 0 for s in Severity}
        for f in findings:
            counts[f.severity.value] += 1
        return counts

    def _risk_score(self, counts: Dict[str, int]) -> int:
        return min(100, counts.get("critical", 0) * 25 + counts.get("high", 0) * 15 +
                   counts.get("medium", 0) * 5 + counts.get("low", 0) * 2)

    def _gen_html(self, title: str, findings: List[Finding], metadata: Dict, output_path: str) -> str:
        counts = self._severity_counts(findings)
        risk = self._risk_score(counts)
        risk_label = "CRITICAL" if risk >= 75 else "HIGH" if risk >= 50 else "MEDIUM" if risk >= 25 else "LOW"

        severity_colors = {
            "critical": "#dc2626", "high": "#ea580c",
            "medium": "#d97706", "low": "#16a34a", "info": "#2563eb",
        }

        def badge(sev: str) -> str:
            color = severity_colors.get(sev, "#888")
            return f'<span style="background:{color};color:white;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold;">{sev.upper()}</span>'

        findings_html = ""
        for f in findings:
            mitre_html = f'<br><small style="color:#888">MITRE: {f.mitre}</small>' if f.mitre else ""
            findings_html += f"""
<div style="border:1px solid #333;border-left:4px solid {severity_colors.get(f.severity.value,'#888')};
            background:#111;margin:10px 0;padding:14px;border-radius:4px;">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;">
    <h3 style="margin:0;color:#e0e0e0;font-size:14px;">{f.title}</h3>
    {badge(f.severity.value)}
  </div>
  <p style="color:#888;margin:6px 0 4px;font-size:12px;">Target: <code style="color:#4cc9f0">{f.target}</code> | Category: {f.category}{mitre_html}</p>
  <p style="color:#ccc;margin:6px 0;font-size:13px;">{f.description[:300]}</p>
</div>"""

        html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>{title}</title>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0a0a0a;color:#e0e0e0;margin:0;padding:0;}}
.header{{background:linear-gradient(135deg,#1a0a0a,#2d0000);padding:40px;border-bottom:3px solid #dc2626;}}
.container{{max-width:1100px;margin:0 auto;padding:30px;}}
.stat-grid{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin:20px 0;}}
.stat{{background:#111;border:1px solid #333;border-radius:8px;padding:14px;text-align:center;}}
.stat-value{{font-size:28px;font-weight:bold;color:#e63946;}}
.stat-label{{font-size:11px;color:#888;margin-top:4px;}}
code{{background:#1a1a1a;padding:2px 5px;border-radius:3px;color:#4cc9f0;font-size:12px;}}
h2{{color:#e63946;border-bottom:1px solid #333;padding-bottom:8px;margin-top:30px;}}
</style></head><body>
<div class="header">
  <h1 style="margin:0;color:#e63946;font-size:28px;">⚔ {title}</h1>
  <p style="color:#888;margin:8px 0 0;">Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")} |
     Operator: {metadata.get('operator','unknown')} |
     Workspace: {metadata.get('workspace','default')}</p>
</div>
<div class="container">
<h2>Executive Summary</h2>
<div class="stat-grid">
  <div class="stat"><div class="stat-value">{len(findings)}</div><div class="stat-label">Total Findings</div></div>
  <div class="stat"><div class="stat-value" style="color:#dc2626">{counts.get('critical',0)}</div><div class="stat-label">Critical</div></div>
  <div class="stat"><div class="stat-value" style="color:#ea580c">{counts.get('high',0)}</div><div class="stat-label">High</div></div>
  <div class="stat"><div class="stat-value" style="color:#d97706">{counts.get('medium',0)}</div><div class="stat-label">Medium</div></div>
  <div class="stat"><div class="stat-value" style="color:#16a34a">{counts.get('low',0)}</div><div class="stat-label">Low</div></div>
  <div class="stat"><div class="stat-value" style="color:#dc2626">{risk}/100</div><div class="stat-label">Risk Score</div></div>
</div>
<p><strong style="color:{'#dc2626' if risk_label in ('CRITICAL','HIGH') else '#d97706'}">Overall Risk: {risk_label}</strong></p>
<h2>Findings ({len(findings)})</h2>
{findings_html}
</div>
</body></html>"""
        Path(output_path).write_text(html, encoding="utf-8")
        log.info(f"HTML report → {output_path}")
        return output_path

    def _gen_markdown(self, title: str, findings: List[Finding], metadata: Dict, output_path: str) -> str:
        counts = self._severity_counts(findings)
        lines = [
            f"# {title}",
            f"",
            f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  ",
            f"**Operator:** {metadata.get('operator', 'unknown')}  ",
            f"**Workspace:** {metadata.get('workspace', 'default')}  ",
            f"",
            f"## Executive Summary",
            f"",
            f"| Severity | Count |",
            f"|----------|-------|",
        ]
        for sev in Severity:
            lines.append(f"| {sev.value.upper()} | {counts.get(sev.value, 0)} |")
        lines += ["", "## Findings", ""]
        for f in findings:
            lines.append(f"### [{f.severity.value.upper()}] {f.title}")
            lines.append(f"**Target:** `{f.target}` | **Category:** {f.category}")
            if f.mitre:
                lines.append(f"**MITRE:** {f.mitre}")
            lines.append(f"")
            lines.append(f.description)
            lines.append("")
        Path(output_path).write_text("\n".join(lines), encoding="utf-8")
        return output_path

    def _gen_json(self, title: str, findings: List[Finding], metadata: Dict, output_path: str) -> str:
        data = {
            "title": title,
            "generated_at": datetime.utcnow().isoformat(),
            "metadata": metadata,
            "summary": self._severity_counts(findings),
            "risk_score": self._risk_score(self._severity_counts(findings)),
            "findings": [
                {
                    "title": f.title,
                    "target": f.target,
                    "severity": f.severity.value,
                    "category": f.category,
                    "description": f.description,
                    "tags": f.tags,
                    "mitre": f.mitre,
                    "cvss": f.cvss,
                    "cve": f.cve,
                }
                for f in findings
            ],
        }
        Path(output_path).write_text(json.dumps(data, indent=2), encoding="utf-8")
        return output_path

    def _gen_xlsx(self, title: str, findings: List[Finding], metadata: Dict, output_path: str) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Findings"
            # Header
            headers = ["Severity", "Title", "Target", "Category", "Description", "MITRE", "CVE", "Tags"]
            sev_fills = {
                "critical": "DC2626", "high": "EA580C",
                "medium": "D97706", "low": "16A34A", "info": "2563EB",
            }
            header_fill = PatternFill(fgColor="1A1A2E", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            # Data
            for row, f in enumerate(findings, 2):
                sev_color = sev_fills.get(f.severity.value, "888888")
                sev_fill = PatternFill(fgColor=sev_color, fill_type="solid")
                ws.cell(row=row, column=1, value=f.severity.value.upper()).fill = sev_fill
                ws.cell(row=row, column=2, value=f.title)
                ws.cell(row=row, column=3, value=f.target)
                ws.cell(row=row, column=4, value=f.category)
                ws.cell(row=row, column=5, value=f.description[:200])
                ws.cell(row=row, column=6, value=f.mitre or "")
                ws.cell(row=row, column=7, value=f.cve or "")
                ws.cell(row=row, column=8, value=",".join(f.tags))
            # Summary sheet
            ws2 = wb.create_sheet("Summary")
            ws2["A1"] = title
            ws2["A1"].font = Font(bold=True, size=14, color="DC2626")
            row = 3
            counts = self._severity_counts(findings)
            for sev, count in counts.items():
                ws2.cell(row=row, column=1, value=sev.upper()).font = Font(bold=True)
                ws2.cell(row=row, column=2, value=count)
                row += 1
            wb.save(output_path)
            return output_path
        except ImportError:
            return self._gen_json(title, findings, metadata, output_path.replace(".xlsx", ".json"))

    def _gen_pdf(self, title: str, findings: List[Finding], metadata: Dict, output_path: str) -> str:
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import inch
            doc = SimpleDocTemplate(output_path, pagesize=letter, topMargin=0.75*inch)
            styles = getSampleStyleSheet()
            story = []
            # Title
            title_style = ParagraphStyle("rtf_title", parent=styles["Title"], textColor=colors.HexColor("#DC2626"), fontSize=18)
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2*inch))
            # Meta
            story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Operator: {metadata.get('operator','unknown')}", styles["Normal"]))
            story.append(Spacer(1, 0.3*inch))
            # Summary table
            counts = self._severity_counts(findings)
            risk = self._risk_score(counts)
            summary_data = [["Severity", "Count"]] + [[k.upper(), str(v)] for k, v in counts.items()]
            t = Table(summary_data, colWidths=[2*inch, 1*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(f"Risk Score: {risk}/100", styles["Heading2"]))
            story.append(Spacer(1, 0.2*inch))
            # Findings
            story.append(Paragraph("Findings", styles["Heading1"]))
            sev_colors_rl = {
                "critical": colors.HexColor("#DC2626"), "high": colors.HexColor("#EA580C"),
                "medium": colors.HexColor("#D97706"), "low": colors.HexColor("#16A34A"),
                "info": colors.HexColor("#2563EB"),
            }
            for f in findings[:200]:
                sev_style = ParagraphStyle(
                    f"sev_{f.severity.value}",
                    parent=styles["Normal"],
                    textColor=sev_colors_rl.get(f.severity.value, colors.black),
                    fontName="Helvetica-Bold",
                    fontSize=11,
                )
                story.append(Paragraph(f"[{f.severity.value.upper()}] {f.title}", sev_style))
                story.append(Paragraph(f"Target: {f.target} | Category: {f.category}", styles["Normal"]))
                if f.mitre:
                    story.append(Paragraph(f"MITRE: {f.mitre}", styles["Italic"]))
                story.append(Paragraph(f.description[:400], styles["Normal"]))
                story.append(Spacer(1, 0.1*inch))
            doc.build(story)
            return output_path
        except ImportError:
            return self._gen_html(title, findings, metadata, output_path.replace(".pdf", ".html"))


# Singleton
report_engine = ReportEngine()
